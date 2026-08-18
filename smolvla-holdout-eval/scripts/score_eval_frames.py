"""Keyboard-score Fixed-tab trials from extracted grasp/late frames into a scoresheet.

Keys per trial:
  y  grasped Y + placed Y (success)
  g  grasped Y + placed N  (then pick failure: d=dropped, w=wrong-placement, o=other)
  n  grasped N + placed N  (failure mode: no-grasp)
  s  skip (leave blank)
  b  go back one trial
  q  quit (saves what is filled)

Usage:
  python scripts/score_eval_frames.py
  python scripts/score_eval_frames.py --start 10
"""

from __future__ import annotations

import argparse
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook

FIRST_TRIAL_ROW = 15


def side_by_side(grasp: np.ndarray, late: np.ndarray, trial: int) -> np.ndarray:
    h = max(grasp.shape[0], late.shape[0])
    w = grasp.shape[1] + late.shape[1]
    canvas = np.zeros((h + 40, w, 3), dtype=np.uint8)
    canvas[40 : 40 + grasp.shape[0], : grasp.shape[1]] = grasp
    canvas[40 : 40 + late.shape[0], grasp.shape[1] :] = late
    cv2.putText(
        canvas,
        f"Trial {trial}   y=success  g=grasped-fail  n=no-grasp  s=skip  b=back  q=quit",
        (8, 28),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.55,
        (255, 255, 255),
        1,
        cv2.LINE_AA,
    )
    cv2.putText(canvas, "GRASP", (8, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
    cv2.putText(
        canvas,
        "LATE",
        (grasp.shape[1] + 8, 60),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.5,
        (0, 255, 0),
        1,
    )
    return canvas


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default="SmolVLA_eval_scoresheet.xlsx")
    ap.add_argument(
        "--frames",
        default="outputs/eval_score_frames/smolvla_fixed_key",
        help="Dir with epXX_grasp.jpg / epXX_late.jpg",
    )
    ap.add_argument("--start", type=int, default=1, help="1-based trial to start at")
    ap.add_argument("--tab", default="Fixed")
    args = ap.parse_args()

    frames = Path(args.frames)
    wb = load_workbook(args.xlsx)
    ws = wb[args.tab]

    i = args.start - 1  # 0-based episode
    while 0 <= i < 50:
        trial = i + 1
        row = FIRST_TRIAL_ROW + i
        g = cv2.imread(str(frames / f"ep{i:02d}_grasp.jpg"))
        late = cv2.imread(str(frames / f"ep{i:02d}_late.jpg"))
        if g is None or late is None:
            raise SystemExit(f"Missing frames for episode {i} under {frames}")
        view = side_by_side(g, late, trial)
        cv2.imshow("score Fixed", view)
        key = cv2.waitKey(0) & 0xFF

        if key in (ord("q"), 27):
            break
        if key == ord("b"):
            i = max(0, i - 1)
            continue
        if key == ord("s"):
            i += 1
            continue
        if key == ord("y"):
            ws[f"C{row}"] = "Y"
            ws[f"D{row}"] = "Y"
            ws[f"E{row}"] = None
        elif key == ord("n"):
            ws[f"C{row}"] = "N"
            ws[f"D{row}"] = "N"
            ws[f"E{row}"] = "no-grasp"
        elif key == ord("g"):
            ws[f"C{row}"] = "Y"
            ws[f"D{row}"] = "N"
            cv2.putText(
                view,
                "failure: d=dropped  w=wrong-placement  o=other",
                (8, view.shape[0] - 12),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.55,
                (0, 255, 255),
                1,
            )
            cv2.imshow("score Fixed", view)
            fk = cv2.waitKey(0) & 0xFF
            ws[f"E{row}"] = {
                ord("d"): "grasped-dropped",
                ord("w"): "wrong-placement",
            }.get(fk, "other")
        else:
            continue  # ignore unknown key

        wb.save(args.xlsx)
        print(f"T{trial}: C={ws[f'C{row}'].value} D={ws[f'D{row}'].value} E={ws[f'E{row}'].value}")
        i += 1

    wb.save(args.xlsx)
    cv2.destroyAllWindows()
    scored = sum(
        1
        for r in range(FIRST_TRIAL_ROW, FIRST_TRIAL_ROW + 50)
        if ws[f"C{r}"].value not in (None, "")
    )
    print(f"Saved {args.xlsx} ({scored}/50 Fixed trials scored).")


if __name__ == "__main__":
    main()
