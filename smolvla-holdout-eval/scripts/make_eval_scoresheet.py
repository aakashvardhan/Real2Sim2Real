"""Generate a real-robot evaluation scoresheet workbook for a policy.

Writes a two-tab xlsx (Fixed + Random) with the same layout as
ACT_eval_scoresheet.xlsx, so `measure_placement_error.py` and
`log_eval_to_wandb.py` work against it unchanged: header on row 14, trials on
rows 15-64, auto-computed summary in rows 4-12.

The Random tab's 50 cube positions are copied from an existing scoresheet
(default: the ACT one) so both policies are evaluated at the identical
positions already marked on the table.

Usage:
    python scripts/make_eval_scoresheet.py                    # -> SmolVLA_eval_scoresheet.xlsx
    python scripts/make_eval_scoresheet.py --policy ACT --out ACT_v3_eval_scoresheet.xlsx
"""

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

FIRST_TRIAL_ROW = 15
LAST_TRIAL_ROW = 64
N_TRIALS = LAST_TRIAL_ROW - FIRST_TRIAL_ROW + 1

# Per-tab layout. The Random tab carries an extra "Gripper Pos" column, which
# shifts every scoring column one to the right.
TABS = {
    "Fixed": {
        "title": "Fixed Position",
        "headers": ["Trial", "Cube pos", "Grasped? (Y/N)", "Placed? (Y/N)",
                    "Failure mode", "Placement err (cm)"],
        "summary_col": "D",
        "grasped": "C",
        "placed": "D",
        "failure": "E",
        "err": "F",
    },
    "Random": {
        "title": "Random Position",
        "headers": ["Trial", "Cube pos (x,y cm)", "Gripper Pos (cm)", "Grasped? (Y/N)",
                    "Placed? (Y/N)", "Failure mode", "Placement err (cm)"],
        "summary_col": "E",
        "grasped": "D",
        "placed": "E",
        "failure": "F",
        "err": "G",
    },
}

FAILURE_MODES = ("no-grasp", "grasped-dropped", "wrong-placement")


def read_random_positions(xlsx_path):
    """Return the 50 `(x, y)` cube-position strings from a scoresheet's Random tab."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "Random" not in wb.sheetnames:
        raise ValueError(f"{xlsx_path} has no 'Random' tab (has {wb.sheetnames})")
    ws = wb["Random"]
    positions = [ws[f"B{row}"].value for row in range(FIRST_TRIAL_ROW, LAST_TRIAL_ROW + 1)]
    missing = [i + 1 for i, p in enumerate(positions) if p in (None, "")]
    if missing:
        raise ValueError(f"Random!B is empty for trial(s) {missing} in {xlsx_path}")
    return positions


def _summary_formulas(cols):
    """Return the summary rows as (label, formula) pairs for one tab's columns."""
    def rng(col):
        return f"${col}${FIRST_TRIAL_ROW}:${col}${LAST_TRIAL_ROW}"

    grasped, placed = rng(cols["grasped"]), rng(cols["placed"])
    failure, err = rng(cols["failure"]), rng(cols["err"])
    both_y = f'(UPPER({grasped})="Y")*(UPPER({placed})="Y")'
    total, successes = f"${cols['summary_col']}$4", f"${cols['summary_col']}$5"
    return [
        ("Total trials", f"=COUNTA({grasped})"),
        ("Successes (grasp Y AND target Y)", f"=SUMPRODUCT({both_y})"),
        ("Success rate", f"=IF({total}=0,0,{successes}/{total})"),
        ("Grasp rate (grasped Y / total)",
         f'=IF({total}=0,0,COUNTIF({grasped},"Y")/{total})'),
        ("Mean placement error (cm, successes only)",
         f'=IFERROR(AVERAGEIFS({err},{grasped},"Y",{placed},"Y"),"-")'),
        ("Placement MSE (cm^2, optional)",
         f'=IFERROR(SUMPRODUCT({both_y}*{err}^2)/{successes},"-")'),
        *[(f"Fail: {mode}", f'=COUNTIF({failure},"{mode}")') for mode in FAILURE_MODES],
    ]


def build_tab(ws, policy, cols, positions):
    ws["A1"] = f"{policy} Evaluation - {cols['title']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "SUMMARY (auto-computed)"
    ws["A3"].font = Font(bold=True)

    for offset, (label, formula) in enumerate(_summary_formulas(cols)):
        row = 4 + offset
        ws[f"A{row}"] = label
        ws[f"{cols['summary_col']}{row}"] = formula

    for i, header in enumerate(cols["headers"]):
        cell = ws.cell(row=14, column=i + 1, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i in range(N_TRIALS):
        row = FIRST_TRIAL_ROW + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = positions[i]

    ws.column_dimensions["A"].width = 40  # holds the summary labels
    for i in range(1, len(cols["headers"])):
        ws.column_dimensions[chr(ord("A") + i)].width = 18
    ws.freeze_panes = "A15"


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--policy", default="SmolVLA", help="Policy name used in the tab titles.")
    ap.add_argument("--out", default=None,
                    help="Output xlsx (default: <policy>_eval_scoresheet.xlsx).")
    ap.add_argument("--positions-from", default="ACT_eval_scoresheet.xlsx",
                    help="Scoresheet to copy the 50 Random cube positions from.")
    ap.add_argument("--force", action="store_true",
                    help="Overwrite the output file if it already exists.")
    args = ap.parse_args()

    out = Path(args.out or f"{args.policy}_eval_scoresheet.xlsx")
    if out.exists() and not args.force:
        raise SystemExit(f"{out} already exists (pass --force to overwrite scored data).")

    random_positions = read_random_positions(args.positions_from)

    wb = Workbook()
    wb.remove(wb.active)
    for name, cols in TABS.items():
        positions = random_positions if name == "Random" else ["(0,0)"] * N_TRIALS
        build_tab(wb.create_sheet(name), args.policy, cols, positions)
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  Fixed  : {N_TRIALS} trials at (0,0)")
    print(f"  Random : {N_TRIALS} trials, positions copied from {args.positions_from}")


if __name__ == "__main__":
    main()
